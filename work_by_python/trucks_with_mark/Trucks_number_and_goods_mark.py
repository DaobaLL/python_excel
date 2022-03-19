from multiprocessing.sharedctypes import Value
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import openpyxl


wb = load_workbook('1647321.xlsx')
ws = wb.active


# 车牌号列表
trucks_number = []

# 车牌号所在单元格位置
position_of_truck_number = []

# 所有卡车所装载的所有货物
loaded_of_trucks = {}


# 获取所有卡车车号
for cell in range(4,ws.max_row):

    # 获取所有 [Tractor No:] 所在的单元格
    if ws['B' + str(cell)].value == "Tractor No:":

        # 将B列对应的卡车号存储到 trucks_number中
        trucks_number.append(ws['C' + str(cell)].value)

        # 将车牌号所在单元格位置存储到列表中
        position_of_truck_number.append('C'+str(cell))

        # 将单个卡车所装载货物信息填入goos_information_of_trucks
        for goods in range(cell+3, ws.max_row):
            if ws['A' + str(goods)].value == '' or ws['A' + str(goods)].value == None:
                break

            # 将卡车上所装在的货物信息填入 loaded_of_trucks 中
            if ws['A' + str(goods)].value == "PKG NO":
                # 单个卡车所装载货物 
                goods_information_of_trucks = []
                for goods_information in range(goods, ws.max_row):

                    if ws['A' + str(goods_information+1)].value == '' or ws['A' + str(goods_information+1)].value == None:
                        break

                    goods_of_truck = ws['A' + str(goods_information+1)].value

                    goods_information_of_trucks.append(goods_of_truck)


                loaded_of_trucks[ws['C' + str(cell)].value] = goods_information_of_trucks


# 将结果写出到 test_out.xlsx
wb2 = load_workbook('trucks_and_mark.xlsx')
ws2 = wb2.active

n=1
# 遍历 loaded_of_trucks
for key,value_1 in loaded_of_trucks.items():

    # 将 loaded_of_trucks 中的车牌写入目标文件的A列
    ws2['A' + str(n)].value = key
    col_out = 2
    for cell_out in value_1:

        if cell_out == "" or cell_out == None:
            break
            
        # 将每辆车所属的货物唛头写入所属车辆之后
        ws2[get_column_letter(col_out) + str(n)].value = cell_out
        col_out += 1
    
    n = n + 1


# 创建卡车车头号与单一唛头对应的字典
shipping_mark_and_truck_number = []

# 将货物唛头与所装载卡车对应
value_1 = None
for key, value_1 in loaded_of_trucks.items():
    
    # 循环提取value_1中的货物唛头
    for good_information in value_1:
        shipping_mark_and_truck_number.append([key, good_information])
        

# 将「车头号与货物」信息写入sheet中
ws3 = wb2['truck_and_mark']

n = 1
for a_col, b_col in shipping_mark_and_truck_number:
    ws3['A' + str(n)].value = a_col
    ws3['B' + str(n)].value = b_col
    n += 1

wb2.save('trucks_and_mark.xlsx')
wb.close
wb2.close