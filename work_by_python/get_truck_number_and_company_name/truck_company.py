import os
import openpyxl
from openpyxl import Workbook, load_workbook

# 读取传入地址的第一级文件夹名称，并存入列表中
def traversalDir_FirstDir(path):
    # 定义一个列表，用来存储结果
    list = []
    # 判断路径是否存在
    if (os.path.exists(path)):
        # 获取该目录下的所有文件或文件夹目录
        files = os.listdir(path)
        for file in files:
            # 得到该文件下所有目录的路径
            m = os.path.join(path, file)
            # 判断该路径下是否是文件夹
            if (os.path.isdir(m)):
                h = os.path.split(m)
                list.append(h[1])
        return list
 
# 传入指定的存放xlsx文件的地址和需要储存到xlsx文件中的列表
def save_list_xlsx(xlsx_path_and_name, directorys_name):


    # 将读取的文件名储存到xlsx中
    wb = Workbook()
    ws = wb.active

    # 遍历整个directorys_name列表
    n = 1
    for dir_name in directorys_name:
        ws['A' + str(n)].value = dir_name
        n += 1

    wb.save(str(xlsx_path_and_name))
    wb.close()

    return print("The result is already saved in file.xlsx")

# 处理提取的文件名，将其分为卡车序号，运输车队，车牌号，提单号
def dirname_to_trucks_informations_list(dirname_list):

    # 所有经过拆分的数据
    trucks_informations = []

    # 单个经过拆分的数据
    truck_information = []

    for single_dir_inf in dirname_list:

        # 卡车号
        truck_number = None

        # 获得开头的卡车号
        for single_letter in single_dir_inf:
            if single_dir_inf.count('-') == 0:
                break

            elif single_letter == '-':
                
                break
            else:
                truck_number = str(truck_number) + str(single_letter)

        truck_information.append(truck_number)

        # 获得车队名称
        single_letter = None
        n = 0
        single_dir_inf_2 = single_dir_inf

        for single_letter in single_dir_inf_2:
            # 如果遍历的字符不是 - 则计数器加一
            if single_letter != '-':
                n += 1

            # 如果single_dir_inf_2 中没有 - 则退出循环
            elif single_dir_inf.count('-') == 0:
                break

            # 如果遍历到 - 则删除提取到的卡车序号部分并将公司名写入卡车信息列表
            else:
                del single_dir_inf_2[0:(n+1)]
                for letter in single_dir_inf_2:
                    if letter == '-':
                        break
                    company_name = str(company_name) + str(letter)
                break
            truck_information.append(company_name)
            



    return trucks_informations



# list_name = traversalDir_FirstDir(r'\\10.168.1.168\业务文档\集装箱\上海公司---客户\8.洛钼-合同过期\A-TFM-2021年-上海\KFMTFM散货包船\第二船\清关境外\装车')

# save_list_xlsx(r'C:\Users\eico\Desktop\work_by_python\get_truck_number_and_company_name\trucks_name.xlsx', list_name)
a = 1
for i in range(1,19):
    a += 1

print(i)
