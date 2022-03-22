from cmath import nan
from operator import index
from openpyxl import load_workbook, Workbook
import numpy as np
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
import time


global index_xlsx
index_xlsx = 1
# 一次性函数，复制取消的合并单元格
def copy_imformation_by_head_cell():
    wb = load_workbook('总箱单表-12点00修改.xlsx')
    ws = wb['总箱单']

    f_colum = ws['F']

    f_cells = []
    for x in range(len(f_colum)):
        if f_colum[x].value == None or f_colum[x].value == '':
            f_colum[x].value = f_colum[x-1].value
        f_cells.append(f_colum[x].value)

    myvar = pd.Series(f_cells)
    print(myvar.head(10))

    # 写入读取的f_cells数据
    n=1
    for cell in f_cells:
        ws['F' + str(n)].value = cell
        n += 1
        
    wb.save('总箱单表-12点00修改.xlsx')
    wb.close

# 读取总的箱单数据，返回经过排序的df
def read_infor():
    
    infor_by_pd = pd.read_excel('总箱单表-12点00修改.xlsx','总箱单')
    # infor_by_pd = pd.read_csv('总箱单表-12点00修改.csv',encoding='gbk')

    # 按照车头号的值进行排序
    sorted_df = infor_by_pd.sort_values(by='提单号',ascending=False)
    sorted_df = sorted_df.sort_values(by='车头号',ascending=True).reset_index()
    return sorted_df
    

# 按照提单号和卡车号排序并出CD申请所需要列的列表
def to_sort_xlsx():
    sorted_df = read_infor()
    # print(sorted_df.head(50))
    # print(sorted_df.tail(50))
    # print(len(sorted_df))
    # print(type(sorted_df.loc[2000,'车头号']))

    # 用来储存一个CD单证所需卡车信息的index
    TheIndexOf_OnTrucksGoodsInformation = []
    
    # 单证所需卡车index的列表
    list_of_all_CD_that_needed_index =[]

    wb = Workbook()
    ws = wb.active
    write_title(ws)
    # 遍历整个表格中的数据，输出{车头号+提单尾号：[信息]} 
    for n in range(0,len(sorted_df)):
        
        # 如果车头号为空，证明货物还没有装车，所以退出循环
        if pd.isna(sorted_df.loc[n,'车头号']) == True:
            break
        
        # 如果上一行车头号、提单号和下一行均相同，则证明这两列的货物装在同一辆车上
        elif sorted_df.loc[n,'车头号'] == sorted_df.loc[n+1, '车头号']:
            if sorted_df.loc[n,'提单号'] == sorted_df.loc[n+1,'提单号']:
                # 车牌号，提单号都相同，继续读取
                TheIndexOf_OnTrucksGoodsInformation.append(n)
            else:
                # 提单号不同，一张CD完成
                TheIndexOf_OnTrucksGoodsInformation.append(n)
                # 出单，调用出单函数
                list_of_all_CD_that_needed_index.append(TheIndexOf_OnTrucksGoodsInformation)
                putout_cd_application(TheIndexOf_OnTrucksGoodsInformation, ws)
                TheIndexOf_OnTrucksGoodsInformation = []
        else:
            # 车牌号不同，一张CD完成
            TheIndexOf_OnTrucksGoodsInformation.append(n)
            # 出单，调用出单函数
            list_of_all_CD_that_needed_index.append(TheIndexOf_OnTrucksGoodsInformation)
            putout_cd_application(TheIndexOf_OnTrucksGoodsInformation, ws)
            TheIndexOf_OnTrucksGoodsInformation = []
    
    wb.save('created_cd_application.xlsx')
    wb.close()
    print('You\'ve done your works, go home NOW!!!!!')
    return list_of_all_CD_that_needed_index


# 出单函数
# 写入表头，执行一次即可
def write_title(ws):
    # 写入表头
    ws['A' + str(1)]='提单号'
    ws['B' + str(1)]='批次号  Column2'
    ws['C' + str(1)]='车头号'
    ws['D' + str(1)]='产品中文名称  Goods Name CN'
    ws['E' + str(1)]='产品英文名称   Goods Name EN'
    ws['F' + str(1)]='毛重'
    ws['G' + str(1)]='商品编码   Hs code.'
    ws['H' + str(1)]='数量'
    ws['I' + str(1)]='CIF 总价   Total CIF PRICE (USD)'
    ws['J' + str(1)]='FOB 总价  Total FOB PRICE  (USD)'
    ws['K' + str(1)]='保险   ISSURANCE (USD)'
    ws['L' + str(1)]='商业发票号   commercial inv no.'
    ws['L' + str(1)]='Purchase order  PO'
    return None

    
def putout_cd_application(TheIndexOf_OnTrucksGoodsInformation, ws):
    #  声明全局变量
    global index_xlsx


    sorted_df = read_infor()
    # 遍历传入的一张cd所需的列表,
    for n in TheIndexOf_OnTrucksGoodsInformation:
        # print(sorted_df.loc[n,]) 
        ws['A' + str(index_xlsx+1)]=sorted_df.loc[n,'提单号']
        ws['B' + str(index_xlsx+1)]=sorted_df.loc[n,'批次号  Column2']
        ws['C' + str(index_xlsx+1)]=sorted_df.loc[n,'车头号']
        ws['D' + str(index_xlsx+1)]=sorted_df.loc[n,'产品中文名称  Goods Name CN']
        ws['E' + str(index_xlsx+1)]=sorted_df.loc[n,'产品英文名称   Goods Name EN']
        ws['F' + str(index_xlsx+1)]=sorted_df.loc[n,'毛重']
        ws['G' + str(index_xlsx+1)]=sorted_df.loc[n,'商品编码   Hs code.']
        ws['H' + str(index_xlsx+1)]=sorted_df.loc[n,'数量']
        ws['I' + str(index_xlsx+1)]=sorted_df.loc[n,'CIF 总价   Total CIF PRICE (USD)']
        ws['J' + str(index_xlsx+1)]=sorted_df.loc[n,'FOB 总价  Total FOB PRICE  (USD)']
        ws['K' + str(index_xlsx+1)]=sorted_df.loc[n,'保险   ISSURANCE (USD)']
        ws['L' + str(index_xlsx+1)]=sorted_df.loc[n,'商业发票号   commercial inv no.']
        ws['M' + str(index_xlsx+1)]=sorted_df.loc[n,'Purchase order  PO']

        index_xlsx += 1

    
    return None


#计算时间消耗
start = time.time()

to_sort_xlsx()

end = time.time()
print("循环运行时间:%.2f秒"%(end-start))
#output:循环运行时间秒
